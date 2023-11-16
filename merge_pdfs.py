import tkinter as tk
from tkinter import messagebox, filedialog
from pathlib import Path
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
import PyPDF2
import os
import re
import time
import shutil
import openpyxl
from PIL import Image
import pytesseract
import fitz
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
import math


def merge_redcliffe_pdf_files():
    # Prompt user to select input directory
    input_directory = filedialog.askdirectory(title="Select Input Directory")

    if input_directory:
        # Prompt user to select output directory
        output_directory = filedialog.askdirectory(title="Select Output Directory")

        if output_directory:
            pdf_dir = Path(input_directory)
            pdf_output_dir = Path(output_directory)
            pdf_output_dir.mkdir(parents=True, exist_ok=True)

            # List all PDF files in the input directory
            pdf_files = list(pdf_dir.glob("*.pdf"))

            keys = set([str(file).split("\\")[-1].split("_")[0].lower() for file in pdf_files])

            for key in keys:
                xray = None
                optometry = None
                ecg_graph = None
                ecg_report = None
                pft = None
                audiometry = None
                vitals = None

                for file in pdf_files:
                    str_pdf_file = str(file)
                    split_str_pdf_files = str_pdf_file.split("_")[1].lower()
                    try:
                        if split_str_pdf_files.endswith(key):
                            pdf_reader = PdfReader(open(file, "rb"))
                            first_page = pdf_reader.pages[0]
                            first_page_text = first_page.extract_text()

                            if "Study Date" and "Report Date" in first_page_text:
                                xray = file
                            elif "OPTOMETRY" in first_page_text:
                                optometry = file
                            elif "Acquired on:" in first_page_text:
                                ecg_graph = file
                            elif "ECG" in first_page_text:
                                ecg_report = file
                            elif "RECORDERS & MEDICARE SYSTEMS" in first_page_text:
                                pft = file
                            elif "VITALS" in first_page_text:
                                print("Vitals")
                                vitals = file
                            else:
                                audiometry = file


                    except Exception as e:
                        print(f"Error processing file: {file}")
                        print(f"Error details: {str(e)}")

                        # Move the problematic file to the error folder
                        error_folder = pdf_output_dir / "error_pdf"
                        error_folder.mkdir(parents=True, exist_ok=True)
                        move_to_error_folder(file, error_folder)

                # Check if at least one file is available for merging
                if xray or optometry or ecg_graph or ecg_report or pft or audiometry or vitals:
                    merger = PdfMerger()
                    if vitals:
                        merger.append(vitals)
                    if ecg_graph:
                        merger.append(ecg_graph)
                    if ecg_report:
                        merger.append(ecg_report)
                    if xray:
                        merger.append(xray)
                    if pft:
                        merger.append(pft)
                    if audiometry:
                        merger.append(audiometry)
                    if optometry:
                        merger.append(optometry)



                    if len(merger.pages) >= 3 and len(merger.pages) <= 4:
                        merged_pdf_dir = pdf_output_dir / "Between_3_And_4"
                    else:
                        merged_pdf_dir = pdf_output_dir / "Less_Then_3_Or_More_Then_4"

                    base_file_name = (
                        xray.stem.split(".")[0].lower() if xray else
                        optometry.stem.split(".")[0].lower() if optometry else
                        ecg_graph.stem.split(".")[0].lower() if ecg_graph else
                        ecg_report.stem.split(".")[0].lower() if ecg_report else
                        pft.stem.split(".")[0].lower() if pft else
                        audiometry.stem.split(".")[0].lower() if audiometry else
                        vitals.stem.split(".")[0].lower() if vitals else
                        "default_name"
                    )
                    print(base_file_name)

                    merged_file_path = merged_pdf_dir / f"{base_file_name}.pdf"
                    merged_file_path.parent.mkdir(parents=True, exist_ok=True)

                    merger.write(str(merged_file_path))
                    print(f"Merged PDF saved to: {merged_file_path}")

            # Display message box after merging is complete
            total_input_count = len(pdf_files)
            pdf_files_3_or_more = list((pdf_output_dir / "Between_3_And_4").glob("*.pdf"))
            total_pdfs_more_than_3 = len(pdf_files_3_or_more)
            pdf_files_3_or_less = list((pdf_output_dir / "Less_Then_3_Or_More_Then_4").glob("*.pdf"))
            total_pdfs_less_than_3 = len(pdf_files_3_or_less)
            total_count = total_pdfs_more_than_3 + total_pdfs_less_than_3
            tk.messagebox.showinfo("PDF Merger", f"Total {total_input_count} PDF files merged into {total_count} PDF files successfully!")

            # Display message box with Less than 3 Pages if any
            if pdf_files_3_or_less:
                file_list = "\n".join(str(file) for file in pdf_files_3_or_less)
                tk.messagebox.showinfo("Missing Files",
                                       f"Total {total_pdfs_less_than_3} merged PDF files have only less then 3 or more then 4 pages:\n{file_list}")
            else:
                tk.messagebox.showinfo("No Missing Files", "All merged PDF files have 3 or 4 pages.")
        else:
            tk.messagebox.showwarning("Output Directory", "Output directory not selected.")
    else:
        tk.messagebox.showwarning("Input Directory", "Input directory not selected.")

def merge_all():
    # Prompt user to select input directory
    input_directory = filedialog.askdirectory(title="Select Input Directory")

    if input_directory:
        # Prompt user to select output directory
        output_directory = filedialog.askdirectory(title="Select Output Directory")

        if output_directory:
            pdf_dir = Path(input_directory)
            pdf_output_dir = Path(output_directory)
            pdf_output_dir.mkdir(parents=True, exist_ok=True)

            # List all PDF files in the input directory
            pdf_files = list(pdf_dir.glob("*.pdf"))
            if pdf_files:
                # Create a PdfFileMerger object
                merger = PdfMerger()
                for pdf_file in pdf_files:
                    merger.append(pdf_file)
                output_file_path = pdf_output_dir / "merged_file.pdf"

                # Write the merged PDF to the output file
                with open(output_file_path, "wb") as output_file:
                    merger.write(output_file)

                print(f"Merged PDF saved to: {output_file_path}")
                tk.messagebox.showinfo("PDF Merger",f"Total {len(pdf_files)} PDF files merged into one PDF successfully!")
            else:
                tk.messagebox.showinfo("No PDF Files", "No PDF files found in the input directory.")
        else:
            tk.messagebox.showwarning("Output Directory", "Output directory not selected.")
    else:
        tk.messagebox.showwarning("Input Directory", "Input directory not selected.")

def rename_pdf_files():
    input_directory = filedialog.askdirectory(title="Select Input Directory")

    if input_directory:
        # Prompt user to select output directory
        output_directory = filedialog.askdirectory(title="Select Output Directory")

        if output_directory:
            input_dir = Path(input_directory)
            output_dir = Path(output_directory)
            output_dir.mkdir(parents=True, exist_ok=True)

            error_dir = output_dir / "error_files"
            error_dir.mkdir(parents=True, exist_ok=True)

            # List all PDF files in the input directory
            pdf_files = list(input_dir.glob("*.pdf"))

            if pdf_files:
                renamed_count = 0
                error_count = 0

                for pdf_file in pdf_files:
                    try:
                        # Open the PDF file in read mode
                        with open(pdf_file, 'rb') as file:
                            pdf_reader = PyPDF2.PdfReader(file)
                            if len(pdf_reader.pages) > 0:
                                first_page = pdf_reader.pages[0]
                                first_page_text = first_page.extract_text()

                                # Extract the patient ID and patient name from the text
                                # X-RAY
                                if "FINDINGS" and "IMPRESSION" in first_page_text:
                                    patient_id = str(first_page_text).split("Patient ID")[1].split(" ")[1]
                                    patient_name = str(first_page_text).split("Name")[1].split("Date")[0].strip()
                                # PFT
                                elif "RECORDERS & MEDICARE SYSTEMS" in first_page_text:
                                    patient_id = str(first_page_text).split("ID     :")[1].split("Age")[0].strip()
                                    patient_name = str(first_page_text).split("Patient:")[1].split("Refd.By:")[0].strip()
                                # ECG GRAPH
                                elif "Acquired on:" in first_page_text:
                                    patient_id = str(first_page_text).split("Id :")[1].split("Name")[0].strip()
                                    patient_name = str(first_page_text).split("Name :")[1].split("Age")[0].strip()
                                    print(patient_id, patient_name)
                                # ECG REPORT
                                else:
                                    # Handle unrecognized format here
                                    raise ValueError("Unrecognized PDF format")

                                renamed_count += 1
                                new_filename = patient_id + "_" + patient_name
                                new_file_path = output_dir / (new_filename + pdf_file.suffix)
                                shutil.copy2(pdf_file, new_file_path)

                                print(f"File renamed and saved: {pdf_file} -> {new_file_path}")
                    except Exception as e:
                        error_count += 1
                        error_file_path = error_dir / pdf_file.name
                        shutil.copy2(pdf_file, error_file_path)
                        print(f"Error processing file {pdf_file}: {str(e)}")

                messagebox.showinfo("Renaming Complete", f"{renamed_count} PDF files have been renamed.")
                if error_count > 0:
                    messagebox.showwarning("Error Files", f"{error_count} PDF files encountered errors. They are saved in the 'error_files' folder.")
            else:
                messagebox.showwarning("No PDF Files", "No PDF files found in the input directory.")
        else:
            messagebox.showwarning("Output Directory", "Output directory not selected.")
    else:
        messagebox.showwarning("Input Directory", "Input directory not selected.")

def extract_patient_data():
    input_directory = filedialog.askdirectory(title="Select Input Directory")

    if input_directory:
        # Prompt user to select output directory
        output_directory = filedialog.askdirectory(title="Select Output Directory")

        if output_directory:
            input_dir = Path(input_directory)
            output_dir = Path(output_directory)
            output_dir.mkdir(parents=True, exist_ok=True)

            error_dir = output_dir / "error_files"
            error_dir.mkdir(parents=True, exist_ok=True)

            pdf_files = list(input_dir.glob("*.pdf"))
            error_count = 0

            patient_data_ecg = []
            patient_data_pft = []
            patient_data_xray = []

            total_ecg_files = 0
            total_pft_files = 0
            total_xray_files = 0

            excel_file_path_ecg = ""
            excel_file_path_pft = ""
            excel_file_path_xray = ""

            workbook_xray = Workbook()
            sheet_xray = workbook_xray.active
            row_xray = 2

            for pdf_file in pdf_files:
                with open(pdf_file, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)

                    if len(pdf_reader.pages) > 0:
                        first_page = pdf_reader.pages[0]
                        first_page_text = first_page.extract_text()

                        try:
                            # Extract XRAY data
                            if "Study Date" in first_page_text or "Report Date" in first_page_text:
                                patient_id = str(first_page_text).split("Patient ID")[1].split(" ")[1].lower().strip()
                                patient = str(first_page_text).split("Name")[1].split("Date")[0].split(" ")[0].strip().lower()
                                if "patient" in patient:
                                    patient_name = patient.split("patient")[0].strip()
                                else:
                                    patient_name = patient
                                age_data = str(first_page_text).split("Age")[1].split("Yr")[0].strip()
                                if "Days" in age_data:
                                    age = age_data.split("Days")[0]
                                else:
                                    age = age_data
                                gender = str(first_page_text).split("Sex")[1].split("Study Date")[0].strip().lower()
                                findings_data = str(first_page_text).split("IMPRESSION")[1].split("Correlate clinically")[0].split(":")[1].strip()
                                if "Please" in findings_data:
                                    findings_with_dot = findings_data.split("Please")[0]
                                    if "•" in findings_with_dot:
                                        findings = findings_with_dot.split("•")[1].split(".")[0]
                                    else:
                                        findings = findings_with_dot.split(".")[0]
                                else:
                                    findings_with_dot = findings_data
                                    if "•" in findings_with_dot:
                                        findings = findings_with_dot.split("•")[1].split(".")[0]
                                    else:
                                        findings = findings_with_dot.split(".")[0]
                                print(patient_id, patient_name, age, gender, findings)
                                patient_data_xray.append((patient_id, patient_name, age, gender, findings))
                                print(patient_id, patient_name, age, gender, findings)
                                total_xray_files += 1

                            # Extract ECG data
                            elif "Acquired on:" in first_page_text:
                                patient_id = str(first_page_text).split("Id :")[1].split(" ")[1].split("\n")[0]
                                patient_name = str(first_page_text).split("Name :")[1].split("Age :")[0]
                                patient_age = str(first_page_text).split("Age :")[1].split(" ")[1].split("\n")[0]
                                patient_gender = str(first_page_text).split("Gender :")[1].split("|")[0].strip()
                                heart_rate = str(first_page_text).split("HR:")[1].split("R(II):")[0].strip()
                                report_time = str(first_page_text).split("Acquired on:")[1][12:17]
                                print(patient_id, patient_name, patient_age, patient_gender, heart_rate, report_time)
                                patient_data_ecg.append((patient_id, patient_name, patient_age, patient_gender, heart_rate, report_time))
                                total_ecg_files += 1

                            #pft
                            elif "RECORDERS & MEDICARE SYSTEMS" in first_page_text:
                                patient_id = str(first_page_text).split("ID     :")[1].split("Age")[0].strip()
                                patient_name = str(first_page_text).split("Patient:")[1].split("Refd.By:")[0].strip()
                                patient_age = str(first_page_text).split("Age    :")[1].split("Yrs")[0].strip()
                                gender = str(first_page_text).split("Gender   :")[1].split("Smoker")[0].strip()
                                height = str(first_page_text).split("Height :")[1].split("Weight")[0].strip()
                                weight = str(first_page_text).split("Weight :")[1].split("Gender")[0].strip()
                                smoker = str(first_page_text).split("Smoker   :")[1].split("Eth.")[0].strip()
                                patient_data_pft.append((patient_id, patient_name, patient_age, gender, height, weight, smoker))
                                total_pft_files += 1

                        except IndexError as e:
                            error_count += 1
                            error_file_path = error_dir / pdf_file.name
                            shutil.copy2(pdf_file, error_file_path)
                            print(f"Error processing file {pdf_file}: Invalid PDF Format")

            if total_ecg_files > 0:
                workbook_ecg = openpyxl.Workbook()
                sheet_ecg = workbook_ecg.active

                sheet_ecg['A1'] = 'patient_id'
                sheet_ecg['B1'] = 'name'
                sheet_ecg['C1'] = 'age'
                sheet_ecg['D1'] = 'gender'
                sheet_ecg['E1'] = 'heart_rate'
                sheet_ecg['F1'] = 'report_time'

                for row, data in enumerate(patient_data_ecg, start=2):
                    sheet_ecg.append(data)

                excel_file_path_ecg = os.path.join(output_dir, "patient_data_ecg.xlsx")
                workbook_ecg.save(excel_file_path_ecg)

            if total_pft_files > 0:
                workbook_pft = openpyxl.Workbook()
                sheet_pft = workbook_pft.active

                sheet_pft['A1'] = 'patient_id'
                sheet_pft['B1'] = 'name'
                sheet_pft['C1'] = 'age'
                sheet_pft['D1'] = 'gender'
                sheet_pft['E1'] = 'height'
                sheet_pft['F1'] = 'weight'
                sheet_pft['G1'] = 'smoker'

                for row, data in enumerate(patient_data_pft, start=2):
                    sheet_pft.append(data)

                excel_file_path_pft = os.path.join(output_dir, "patient_data_pft.xlsx")
                workbook_pft.save(excel_file_path_pft)

            if total_xray_files > 0:
                workbook_xray = openpyxl.Workbook()
                sheet_xray = workbook_xray.active

                sheet_xray['A1'] = 'patient_id'
                sheet_xray['B1'] = 'name'
                sheet_xray['C1'] = 'age'
                sheet_xray['D1'] = 'gender'
                sheet_xray['E1'] = 'Findings'

                for row, data in enumerate(patient_data_xray, start=2):
                    sheet_xray.append(data)

                for row in range(2, len(patient_data_xray) + 2):
                    cell = sheet_xray.cell(row=row, column=5)
                    findings = cell.value
                    if "No significant abnormality seen" in findings:
                        cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
                    else:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

                excel_file_path_xray = os.path.join(output_dir, "patient_data_xray.xlsx")
                workbook_xray.save(excel_file_path_xray)

            message = f"Total {total_ecg_files} ECG and {total_pft_files} PFT and {total_xray_files} XRAY data files have been extracted and saved successfully.\n\n"
            message += f"ECG Output File: {excel_file_path_ecg}\n\nPFT Output File: {excel_file_path_pft}\n\nXRAY Output File: {excel_file_path_xray}"
            messagebox.showinfo("Patient Data Extractor", message)

        else:
            messagebox.showwarning("Output Folder Not Selected", "Output folder not selected.")
    else:
        messagebox.showwarning("Input Folder Not Selected", "Input folder not selected.")


def check_pdf_files():
    # Prompt user to select merged PDF folder
    pdf_folder_path = filedialog.askdirectory(title="Select Merged PDF Folder", mustexist=True)

    if pdf_folder_path:
        # Prompt user to select Excel sheet
        excel_file_path = filedialog.askopenfilename(title="Select Excel Sheet", filetypes=[("Excel Files", "*.xlsx;*.xls")])

        if excel_file_path:
            # Create a workbook and add a worksheet
            wb = Workbook()
            ws = wb.active

            # Add headers to the worksheet
            headers = ["patient_id", "patient_name", "age", "gender", "ECG_GRAPH", "ECG_REPORT", "XRAY_REPORT", "XRAY_IMAGE", "PFT",
                       "AUDIOMETRY", "OPTOMETRY", "VITALS", "PROBLEM"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)

            # Load the Excel sheet for comparison
            comparison_df = pd.read_excel(excel_file_path)

            # Iterate through rows in the Excel sheet
            for _, excel_row in comparison_df.iterrows():
                # Get the corresponding PDF file path based on the patient ID
                pdf_id_prefix = str(excel_row['patient_id']).lower()
                pdf_files = []
                extra_pdf_files = []
                # pdf_files = [file.lower() for file in os.listdir(pdf_folder_path) if file.startswith(pdf_id_prefix)]
                for file in os.listdir(pdf_folder_path):
                    if file.lower().startswith(pdf_id_prefix):
                        pdf_files.append(file.lower())

                # Initialize modality matching list outside the PDF page loop
                modality_match_list = []
                problem_list = []
                if not pdf_files:
                    # No matching PDF file found for patient ID
                    problem_list.append("Pdf file is missing")
                    modality_match_list = ["No"] * 8

                    # Write the results to the worksheet
                    row_data = [
                                   str(excel_row["patient_id"]).lower(),
                                   str(excel_row["patient_name"]).split(" ")[0].lower(),
                                   str(excel_row.get("age", "")).strip(),
                                   str(excel_row["gender"]).strip(),
                               ] + modality_match_list + [', '.join(problem_list)]
                    ws.append(row_data)

                    current_row = ws.max_row

                    # Define the column indices for "Yes" and "No" values
                    yes_columns = [5, 6, 7, 8, 9, 10, 11, 12]  # Assuming columns E to L are modality columns

                    # Apply fill color to cells based on "Yes" or "No"
                    for col_num in range(5, 13):  # Columns E to L
                        cell = ws.cell(row=current_row, column=col_num)
                        if cell.value == "Yes":
                            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00",
                                                    fill_type="solid")  # Green color
                        elif cell.value == "No":
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                    fill_type="solid")  # Red color

                if pdf_files:
                    # Take the first matching PDF file for simplicity, you may need to handle multiple files
                    pdf_file = pdf_files[0]
                    pdf_path = os.path.join(pdf_folder_path, pdf_file)

                    # Extract patient data from the Excel row
                    patient_data_excel = {
                        "patient_id": str(excel_row["patient_id"]).lower().strip(),
                        "patient_name": str(excel_row["patient_name"]).split(" ")[0].lower().strip(),
                        "age": str(excel_row.get("age", "")).strip(),
                        "gender": str(excel_row["gender"]).strip().lower()
                    }

                    print(patient_data_excel)

                    # Iterate through the modalities
                    for modality in ["ECG_GRAPH", "ECG_REPORT", "XRAY_REPORT", "XRAY_IMAGE", "PFT", "AUDIOMETRY", "OPTOMETRY", "VITALS"]:
                        # Initialize modality matching flag for the current modality
                        modality_match = False

                        patient_id = None
                        patient_name = None
                        age = None
                        gender = None
                        # Open the PDF file for the current row
                        pdf_reader = PdfReader(open(pdf_path, "rb"))
                        # Iterate through the PDF pages
                        for page_num in range(len(pdf_reader.pages)):
                            page = pdf_reader.pages[page_num]
                            page_text = page.extract_text()

                            # Initialize a list to track missing PDF modalities
                            missing_modalities = []

                            # Extract patient data based on patterns
                            if modality == "ECG_GRAPH" and "Acquired on:" in page_text:
                                if "Id :" in page_text:
                                    patient_id = str(page_text).split("Id :")[1].split(" ")[1].split("\n")[0].strip().lower()
                                elif "Id:" in page_text:
                                    patient_id = str(page_text).split("Id:")[1].split(" ")[1].split("\n")[0].strip().lower()
                                else:
                                    patient_id = 12345

                                if "Name :" in page_text:
                                    patient_name = str(page_text).split("Name :")[1].split("Age")[0].split(" ")[1].strip().lower()
                                elif "Name:" in page_text:
                                    patient_name = str(page_text).split("Name:")[1].split("Age")[0].split(" ")[1].strip().lower()
                                else:
                                    patient_name = 'invalid'

                                if "Age :" in page_text:
                                    age = str(page_text).split("Age :")[1].split(" ")[1].split("\n")[0].strip()
                                elif "Age:" in page_text:
                                    age = str(page_text).split("Age:")[1].split(" ")[1].split("\n")[0].strip()
                                else:
                                    age = '0'

                                if "Gender :" in page_text:
                                    gender = str(page_text).split("Gender :")[1].split("|")[0].strip().lower()
                                elif "Gender:" in page_text:
                                    gender = str(page_text).split("Gender:")[1].split("|")[0].strip().lower()
                                else:
                                    gender = 'invalid'

                                print("ECG GRAPH", patient_id, patient_name, age, gender)

                                # Compare with Excel data
                                if (patient_id == patient_data_excel["patient_id"] and
                                        patient_name == patient_data_excel["patient_name"] and
                                        age == patient_data_excel["age"] and
                                        gender == patient_data_excel["gender"]):
                                    modality_match = True
                                    break
                                else:
                                    if (patient_id != patient_data_excel["patient_id"] and
                                            patient_name != patient_data_excel["patient_name"] and
                                            age != patient_data_excel["age"] and
                                            gender != patient_data_excel["gender"]):
                                        problem_list.append(f' {modality}: All the data are incorrect')

                            elif modality == "ECG_REPORT" and "ECG" in page_text:
                                patient_name = str(page_text).split("Name:")[1].split("Patient")[0]
                                if patient_name.count(" ") == 1:
                                    patient_name = patient_name.strip().lower()
                                else:
                                    patient_name = patient_name.split(" ")[1].lower().strip()

                                patient_id = str(page_text).split("Patient ID:")[1].split("Age")[0].strip().lower()
                                age = str(page_text).split("Age:")[1].split("Gender")[0].strip()
                                gender = str(page_text).split("Gender:")[1].split("Test")[0].strip().lower()
                                print("ECG REPORT", patient_id, patient_name, age, gender)

                                # Compare with Excel data
                                if (patient_id == patient_data_excel["patient_id"] and
                                        patient_name == patient_data_excel["patient_name"] and
                                        age == patient_data_excel["age"] and
                                        gender == patient_data_excel["gender"]):
                                    modality_match = True
                                    break
                                else:
                                    if (patient_id != patient_data_excel["patient_id"] and
                                            patient_name != patient_data_excel["patient_name"] and
                                            age != patient_data_excel["age"] and
                                            gender != patient_data_excel["gender"]):
                                        problem_list.append(f' {modality}: All the data are incorrect')

                            elif modality == "XRAY_REPORT" and "Study Date" and "Report Date" in page_text:
                                patient_id = str(page_text).split("Patient ID")[1].split(" ")[1].lower().strip()
                                patient = str(page_text).split("Name")[1].split("Date")[0].split(" ")[0].strip().lower()
                                if "patient" in patient:
                                    patient_name = patient.split("patient")[0].strip()
                                else:
                                    patient_name = patient
                                age = str(page_text).split("Age")[1].split("Yr")[0].strip()
                                gender = str(page_text).split("Sex")[1].split("Study Date")[0].strip().lower()
                                print('XRAY', patient_id, patient_name, age, gender)

                                # Compare with Excel data
                                if (patient_id == patient_data_excel["patient_id"] and
                                        patient_name == patient_data_excel["patient_name"] and
                                        age == patient_data_excel["age"] and
                                        gender == patient_data_excel["gender"]):
                                    modality_match = True
                                    break
                                else:
                                    if (patient_id != patient_data_excel["patient_id"] and
                                            patient_name != patient_data_excel["patient_name"] and
                                            age != patient_data_excel["age"] and
                                            gender != patient_data_excel["gender"]):
                                        problem_list.append(f' {modality}: All the data are incorrect')

                            elif modality == "XRAY_IMAGE" and "Page 2 of 2" in page_text:
                                if "Page 2 of 2" in page_text:
                                    modality_match = True
                                    break

                            elif modality == "PFT" and "RECORDERS & MEDICARE SYSTEMS" in page_text:
                                patient_name = str(page_text).split("Patient: ")[1].split("Refd.By:")[0].split("\n")[0].lower()
                                if " " in patient_name:
                                    patient_name = patient_name.split(" ")[0]
                                else:
                                    patient_name = patient_name
                                patient_id = str(page_text).split("ID     :")[1].split("Age")[0].strip().lower()
                                age = str(page_text).split("Age    :")[1].split("Yrs")[0].strip()
                                if "Smoker" in page_text:
                                    gender = str(page_text).split("Gender   :")[1].split("Smoker")[0].strip().lower()
                                else:
                                    gender = str(page_text).split("Gender   :")[1].split("Eth. Corr:")[0].strip().lower()

                                print('PFT', patient_id, patient_name, age, gender)

                                # Compare with Excel data
                                if (patient_id == patient_data_excel["patient_id"] and
                                        patient_name == patient_data_excel["patient_name"] and
                                        age == patient_data_excel["age"] and
                                        gender == patient_data_excel["gender"]):
                                    modality_match = True
                                    break
                                else:
                                    if (patient_id != patient_data_excel["patient_id"] and
                                            patient_name != patient_data_excel["patient_name"] and
                                            age != patient_data_excel["age"] and
                                            gender != patient_data_excel["gender"]):
                                        problem_list.append(f' {modality}: All the data are incorrect')

                            elif modality == "AUDIOMETRY" and "AUDIOMETRY" in page_text:
                                patient_name = str(page_text).split("Name:")[1].split("Patient ID:")[0].split(" ")[1].strip().lower()
                                patient_id = str(page_text).split("Patient ID:")[1].split("Age")[0].strip().lower()
                                age = str(page_text).split("Age:")[1].split("Gender")[0].strip()
                                gender = str(page_text).split("Gender:")[1].split("Test")[0].strip().lower()
                                print(patient_id, patient_name, age, gender)

                                # Compare with Excel data
                                if (patient_id == patient_data_excel["patient_id"] and
                                        patient_name == patient_data_excel["patient_name"] and
                                        age == patient_data_excel["age"] and
                                        gender == patient_data_excel["gender"]):
                                    modality_match = True
                                    break
                                else:
                                    if (patient_id != patient_data_excel["patient_id"] and
                                            patient_name != patient_data_excel["patient_name"] and
                                            age != patient_data_excel["age"] and
                                            gender != patient_data_excel["gender"]):
                                        problem_list.append(f' {modality}: All the data are incorrect')

                            elif modality == "OPTOMETRY" and "OPTOMETRY" in page_text:
                                patient_name = str(page_text).split("Name:")[1].split("Patient ID:")[0].split(" ")[1].strip().lower()
                                patient_id = str(page_text).split("Patient ID:")[1].split("Age")[0].strip().lower()
                                age = str(page_text).split("Age:")[1].split("Gender")[0].strip()
                                gender = str(page_text).split("Gender:")[1].split("Test")[0].strip().lower()
                                print('OPTOMETRY', patient_id, patient_name, age, gender)

                                # Compare with Excel data
                                if (patient_id == patient_data_excel["patient_id"] and
                                        patient_name == patient_data_excel["patient_name"] and
                                        age == patient_data_excel["age"] and
                                        gender == patient_data_excel["gender"]):
                                    modality_match = True
                                    break
                                else:
                                    if (patient_id != patient_data_excel["patient_id"] and
                                            patient_name != patient_data_excel["patient_name"] and
                                            age != patient_data_excel["age"] and
                                            gender != patient_data_excel["gender"]):
                                        problem_list.append(f' {modality}: All the data are incorrect')

                            elif modality == "VITALS" and "VITALS" in page_text:
                                patient_name = str(page_text).split("Name:")[1].split("Patient ID:")[0].split(" ")[1].strip().lower()
                                patient_id = str(page_text).split("Patient ID:")[1].split("Age")[0].strip().lower()
                                age = str(page_text).split("Age:")[1].split("Gender")[0].strip()
                                gender = str(page_text).split("Gender:")[1].split("Test")[0].strip().lower()
                                print(patient_id, patient_name, age, gender)

                                # Compare with Excel data
                                if (patient_id == patient_data_excel["patient_id"] and
                                    patient_name == patient_data_excel["patient_name"] and
                                    age == patient_data_excel["age"] and
                                    gender == patient_data_excel["gender"]):

                                    modality_match = True
                                    break
                                else:
                                    if (patient_id != patient_data_excel["patient_id"] and
                                            patient_name != patient_data_excel["patient_name"] and
                                            age != patient_data_excel["age"] and
                                            gender != patient_data_excel["gender"]):
                                        problem_list.append(f' {modality}: All the data are incorrect')

                        issues = []
                        if patient_id == patient_data_excel["patient_id"] or patient_name == patient_data_excel["patient_name"] or age == patient_data_excel["age"] or gender == patient_data_excel["gender"]:
                            if not modality_match:
                                if patient_id != patient_data_excel["patient_id"]:
                                    issues.append("ID")
                                if patient_name != patient_data_excel["patient_name"]:
                                    issues.append("Name")
                                if age != patient_data_excel["age"]:
                                    issues.append("Age")
                                if gender != patient_data_excel["gender"]:
                                    issues.append("Gender")

                        # Append the modality 2 corresponding issues to the problem_list
                        if issues:
                            problem_list.append(f"{modality}: {', '.join(issues)}")
                        modality_match_list.append("Yes" if modality_match else "No")

                    # Write the results to the worksheet
                    row_data = [
                        patient_data_excel["patient_id"],
                        patient_data_excel["patient_name"],
                        patient_data_excel["age"],
                        patient_data_excel["gender"]
                    ] + modality_match_list  + [', '.join(problem_list)]
                    ws.append(row_data)

                    current_row = ws.max_row

                    # Define the column indices for "Yes" and "No" values
                    yes_columns = [5, 6, 7, 8, 9, 10, 11, 12]  # Assuming columns E to L are modality columns

                    # Apply fill color to cells based on "Yes" or "No"
                    for col_num in range(5, 13):  # Columns E to L
                        cell = ws.cell(row=current_row, column=col_num)
                        if cell.value == "Yes":
                            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00",
                                                    fill_type="solid")  # Green color
                        elif cell.value == "No":
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000",
                                                    fill_type="solid")  # Red color
                else:
                    print(f"No matching PDF file found for patient ID: {pdf_id_prefix}")

            output_directory = filedialog.askdirectory(title="Select Output Directory")
            if output_directory:
                output_filename = "patient_data_comparison.xlsx"
                wb.save(os.path.join(output_directory, output_filename))
                print("Data comparison completed.")
            else:
                print("Output directory not selected.")
        else:
            print("Excel sheet not selected.")
    else:
        print("Merged PDF folder not selected.")


def sanitize_filename(filename):
    # Remove invalid characters from the filename
    return re.sub(r'[\\/:*?"<>|]', '_', filename)

def split_patient_file():
    input_directory = filedialog.askdirectory(title="Select Input Directory")

    if input_directory:
        output_directory = filedialog.askdirectory(title="Select Output Directory")

        if output_directory:
            pdf_dir = Path(input_directory)
            pdf_output_dir = Path(output_directory)
            pdf_output_dir.mkdir(parents=True, exist_ok=True)

            pdf_files = list(pdf_dir.glob("*.pdf"))

            if pdf_files:
                for input_pdf_path in pdf_files:
                    try:
                        # Open the merged PDF file
                        with open(input_pdf_path, 'rb') as pdf_file:
                            pdf_reader = PyPDF2.PdfReader(pdf_file)

                            # Create a subdirectory for the patient if it doesn't exist
                            patient_id = sanitize_filename(input_pdf_path.stem)
                            patient_dir = pdf_output_dir / patient_id
                            patient_dir.mkdir(parents=True, exist_ok=True)

                            # Loop through each page in the PDF and save them as individual PDF files
                            for page_number in range(len(pdf_reader.pages)):
                                pdf_writer = PyPDF2.PdfWriter()
                                pdf_writer.add_page(pdf_reader.pages[page_number])

                                # Extract text data from the page to determine the modality
                                page_text = pdf_reader.pages[page_number].extract_text()
                                modality = None

                                if "Study Date" and "Report Date" in page_text:
                                    modality = 'XRAY_REPORT'
                                elif "RECORDERS & MEDICARE SYSTEMS" in page_text:
                                    modality = 'PFT'
                                elif "Page 2 of 2" in page_text:
                                    modality = 'XRAY IMAGE'
                                elif "OPTOMETRY" in page_text:
                                    modality = 'OPTOMETRY'
                                else:
                                    modality = 'AUDIOMETRY'

                                if modality:
                                    output_file_path = patient_dir / f'{patient_id}_{modality}.pdf'
                                else:
                                    output_file_path = patient_dir / f'{patient_id}_page_{page_number + 1}.pdf'

                                # Save the individual page as a PDF file with the new name
                                with open(output_file_path, 'wb') as output_file:
                                    pdf_writer.write(output_file)

                            print(f"PDF files for patient {patient_id} split and renamed successfully.")
                    except Exception as e:
                        print(f"Error processing {input_pdf_path}: {str(e)}")
                        continue  # Skip this file and continue with the next

                print("PDF files processed.")
            else:
                print("No PDF files found in the input directory.")

        else:
            print("Output directory not selected.")
    else:
        print("Input directory not selected.")


# Create the main window
window = tk.Tk()
window.title("Camp - Automation Tools")
# Set the window dimensions and position it on the screen
window.geometry("1000x500+200-100")


redcliffe_label = tk.Label(window, text="Merge Pdf Files", font=("Arial", 16, "bold"))
redcliffe_label.place(x=580, y=10, anchor='ne')

merge_redcliffe_button1 = tk.Button(window, bg='blue', fg='white', activebackground='darkblue', activeforeground='white', padx=30, pady=10, relief='raised', text="Merge PDF Files", command=merge_redcliffe_pdf_files, font=("Arial", 12, "bold"))
merge_redcliffe_button2 = tk.Button(window, bg='magenta', fg='black', activebackground='gold', activeforeground='black', padx=30, pady=10, relief='raised', text="Merge All PDF Files", command=merge_all, font=("Arial", 12, "bold"))
merge_redcliffe_button1.place(x=600, y=58, anchor='ne')
merge_redcliffe_button2.place(x=623, y=178, anchor='ne')

pdf_rename_label = tk.Label(window, text="File Renaming System", font=("Arial", 16, "bold"))
pdf_rename_label.pack(pady=10, padx=20, anchor='w')

pdf_rename_button1 = tk.Button(window, bg='orange', fg='black', activebackground='darkblue', activeforeground='white', padx=30, pady=10, relief='raised', text="Rename PDF Files", command=rename_pdf_files, font=("Arial", 12, "bold"))
pdf_rename_button1.pack(pady=8, padx=20, anchor='w')

generate_excel_label = tk.Label(window, text="Data Extracting System", font=("Arial", 16, "bold"))
generate_excel_label.place(x=259, y=130, anchor='ne')

generate_excel_button = tk.Button(window, bg='pink',fg='black', activebackground='darkblue', activeforeground='white',padx=30, pady=10, relief='raised', text="Generate Patient Excel", command=extract_patient_data, font=("Arial", 12, "bold"))
generate_excel_button.place(x=262, y=180, anchor='ne')

check_pdf_File = tk.Label(window, text="Check Pdf Files", font=("Arial", 16, "bold"))
check_pdf_File.place(x=920, y=10, anchor='ne')

check_pdf_button = tk.Button(window, bg='green',fg='black', activebackground='darkblue', activeforeground='white',padx=30, pady=10, relief='raised', text="Check Pdf Files", command=check_pdf_files, font=("Arial", 12, "bold"))
check_pdf_button.place(x=956, y=57, anchor='ne')

check_pdf_File = tk.Label(window, text="Split Pdf Files", font=("Arial", 16, "bold"))
check_pdf_File.place(x=903, y=130, anchor='ne')

check_pdf_button = tk.Button(window, bg='yellow',fg='black', activebackground='darkblue', activeforeground='white',padx=30, pady=10, relief='raised', text="Split Pdf Files", command=split_patient_file, font=("Arial", 12, "bold"))
check_pdf_button.place(x=940, y=175, anchor='ne')


window.mainloop()






